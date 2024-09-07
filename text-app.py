import json
import time
import requests
from requests.exceptions import RequestException
import logging
import sseclient
import os
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl import load_workbook
import re

load_dotenv()

api_key = os.getenv("API_KEY")
api_host = os.getenv("API_HOST")
api_model = os.getenv("MODEL")

# 设置日志
logging.basicConfig(
    filename="text-log.txt",
    level=logging.INFO,
    format="%(asctime)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    encoding="utf-8"  # 添加这一行来指定编码
)

with open("text-prompts-system.text", "r", encoding="utf-8") as file:
    system_prompt = file.read()


# 请求 URL 和头信息
url = f"{api_host}/v1/chat/completions"

headers = {
    "Content-Type": "application/json",
    "Accept": "text/event-stream",
    "Authorization": f"Bearer {api_key}",
}

# 请求数据模板
data_template = {
    "messages": [
        {"role": "system", "content": ""},
        {"role": "user", "content": ""},
    ],
    "stream": True,
    "model": api_model,
    "temperature": 0.5,
    "presence_penalty": 0,
    "frequency_penalty": 0,
    "top_p": 1,
}


def send_request(prompt):
    data = data_template.copy()
    data["messages"][0]["content"] = system_prompt
    data["messages"][1]["content"] = prompt

    log_message = f"正在创建 '{prompt}' 的对话线程。"
    print(log_message)
    logging.info(log_message)

    try:
        response = requests.post(url, headers=headers, json=data, stream=True)
        client = sseclient.SSEClient(response)
        result = ""
        for event in client.events():
            if event.data == "[DONE]" or event.data == " [DONE]":
                break

            try:
                data = json.loads(event.data)
                if "choices" in data and len(data["choices"]) > 0:
                    content = data["choices"][0]["delta"].get("content")
                    
                    if content:
                        result += content
                        print(content, end="", flush=True)
                    else:
                        print(json.dumps(data["choices"], indent=2))
                else:
                    print(json.dumps(data, indent=2))
            except json.JSONDecodeError:
                print(f"无法解析JSON: {event.data}")
                logging.info(event.data)

        print()  # 打印一个换行

        log_message = f"'{prompt}' 的对话线程已结束。"

        logging.info(result)
        save_to_excel(prompt, result)

        print(log_message)
        logging.info(log_message)
    except RequestException as e:
        error_message = f"请求发生错误: {e}"
        print(error_message)
        logging.error(error_message)
        return False
    return True

def save_to_excel(prompt, log_message):
    current_date = datetime.now().strftime("%Y%m%d")
    output_dir = "output"
    filename = f"{output_dir}/text_{current_date}.xlsx"
    
    # 确保output文件夹存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # 从log_message中提取第一个JSON内容
    json_match = re.search(r'\{.*?\}', log_message, re.DOTALL)
    if json_match:
        json_content = json.loads(json_match.group())
    else:
        json_content = {}

    # 准备表头和数据
    headers = ["Prompt", "Response"] + list(json_content.keys())
    row_data = [prompt, log_message] + list(json_content.values())

    # 保存逻辑
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        if ws[1][0].value != headers[0]:  # 检查表头是否匹配
            ws.insert_rows(1)
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header)
    else:
        wb = Workbook()
        ws = wb.active
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

    # 添加新行
    ws.append(row_data)
    wb.save(filename)

def main():
    # 读取Excel文件
    wb = load_workbook('text-data.xlsx')
    ws = wb.active
    
    # 获取表头
    headers = [cell.value for cell in ws[1]]
    
    # 将每一行转换为JSON字符串
    json_strings = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: value for i, value in enumerate(row) if value is not None}
        json_string = json.dumps(row_dict, ensure_ascii=False)
        json_strings.append(json_string)
    
    # 将JSON字符串列表写入prompts-text.json文件
    with open('text-prompts.json', 'w', encoding='utf-8') as f:
        json.dump(json_strings, f, ensure_ascii=False, indent=4)

    # 读取 JSON 文件
    with open("text-prompts.json", "r", encoding="utf-8") as file:
        prompts = json.load(file)

    # 原有的main函数内容
    for prompt in prompts:
        try:
            if not send_request(prompt):
                continue
            log_message = "等待0.1分钟后继续下一个请求。"
            time.sleep(10)  # 等待 10 秒
            print(log_message)
            logging.info(log_message)
        except RequestException as e:
            error_message = f"请求发生错误: {e}"
            print(error_message)
            logging.info(error_message)

    print("所有的任务都完成了！")
    logging.info("所有的任务都完成了！")

if __name__ == "__main__":
    main()
